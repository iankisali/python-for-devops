import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
product_under_10 = {}

print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # calculation for number of products per supplier
    if supplier_name in products_per_supplier:
        current_no_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_no_products + 1
    else:
        print('Adding new supplier!!')
        products_per_supplier[supplier_name] = 1

    # calculate total number of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # products with inventory under 10
    if inventory < 10:
        product_under_10[int(product_num)] = int(inventory)

    # adding value for total inventory file
    inventory_price.value = inventory * price


print(products_per_supplier)
print(total_value_per_supplier)
print(product_under_10)

inventory_file.save("inventory_total.xlsx")
